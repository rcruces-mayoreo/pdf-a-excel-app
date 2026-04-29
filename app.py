import streamlit as st
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment
import io

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="PDF to Excel Pro", page_icon="📄", layout="wide")

def clean_sheet_name(name, existing_names):
    """Limpia y asegura que el nombre de la hoja sea único y válido."""
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, "")
    name = name[:30] 
    base_name = name
    counter = 1
    while name in existing_names:
        suffix = f"_{counter}"
        name = base_name[:31 - len(suffix)] + suffix
        counter += 1
    return name

def process_pdfs(uploaded_pdfs, ancho_deseado, dpi_calidad):
    """Función principal que procesa los archivos."""
    wb = Workbook()
    wb.remove(wb.active) # Quitar hoja por defecto
    existing_sheets = []
    
    progress_bar = st.progress(0)
    total_files = len(uploaded_pdfs)

    for idx, archivo_pdf in enumerate(uploaded_pdfs):
        sheet_name = clean_sheet_name(archivo_pdf.name.replace(".pdf", ""), existing_sheets)
        existing_sheets.append(sheet_name)
        
        ws = wb.create_sheet(title=sheet_name)
        ws["D2"] = f"ARCHIVO: {archivo_pdf.name}"
        ws["D2"].font = Font(bold=True, size=14)
        
        fila_actual = 4  
        image_counter = 0 
        
        pdf_bytes = archivo_pdf.read()
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        for pagina in doc:
            pix = pagina.get_pixmap(dpi=dpi_calidad)
            img_data = pix.tobytes("png")
            
            image_stream = io.BytesIO(img_data)
            img_excel = ExcelImage(image_stream)
            
            proporcion = ancho_deseado / pix.width
            alto_deseado = int(pix.height * proporcion)
            img_excel.width = ancho_deseado
            img_excel.height = alto_deseado
            
            celda_destino = "A" if image_counter % 2 == 0 else "H"
            ws.add_image(img_excel, f"{celda_destino}{fila_actual}")
            
            if image_counter % 2 != 0:
                filas_necesarias = int(alto_deseado / 18) + 2
                fila_actual += filas_necesarias
            
            image_counter += 1
        
        progress_bar.progress((idx + 1) / total_files)
        
    return wb

# --- INTERFAZ DE USUARIO ---
st.title("📄 PDF to Excel (Custom Name)")

# Nueva sección: Nombre del archivo
st.subheader("1. Configuración del archivo")
col1, col2 = st.columns(2)

with col1:
    # --- AQUÍ ESTÁ LA NUEVA FUNCIÓN ---
    nombre_archivo_usuario = st.text_input(
        "Nombre del archivo Excel de salida:",
        value="auditoria_facturas",
        help="No es necesario escribir .xlsx al final"
    )

with col2:
    # Asegurar que el nombre termine en .xlsx
    if not nombre_archivo_usuario.strip():
        nombre_final = "reporte_sin_nombre.xlsx"
    elif not nombre_archivo_usuario.lower().endswith(".xlsx"):
        nombre_final = f"{nombre_archivo_usuario}.xlsx"
    else:
        nombre_final = nombre_archivo_usuario

# Barra lateral para ajustes técnicos
with st.sidebar:
    st.header("⚙️ Ajustes de Imagen")
    ancho = st.slider("Ancho de imagen", 300, 800, 500)
    calidad = st.select_slider("Calidad (DPI)", options=[72, 100, 150, 200], value=150)

st.subheader("2. Carga de archivos")
uploaded_pdfs = st.file_uploader("Sube tus PDFs", type=['pdf'], accept_multiple_files=True)

if uploaded_pdfs:
    if st.button(f"🚀 Generar {nombre_final}"):
        with st.spinner('Procesando...'):
            try:
                workbook_final = process_pdfs(uploaded_pdfs, ancho, calidad)
                
                excel_buffer = io.BytesIO()
                workbook_final.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.success(f"¡Listo! El archivo '{nombre_final}' se ha generado correctamente.")
                
                st.download_button(
                    label=f"📥 Descargar {nombre_final}",
                    data=excel_buffer,
                    file_name=nombre_final, # Usamos el nombre dinámico aquí
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error inesperado: {e}")
